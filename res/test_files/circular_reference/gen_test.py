import openpyxl
from openpyxl import Workbook

def create_direct_circular_reference():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Create direct circular reference
    cell = ws['A1']
    cell.value = '=A1+1'  # This sets the formula
    cell.data_type = 'f'  # Explicitly set as formula type
    
    # Save with keep_vba=True to preserve formulas
    wb.save('direct_circular.xlsx')
    print("Created direct_circular.xlsx")

def create_indirect_circular_reference():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Create indirect circular reference chain
    ws['A1'].value = '=B1+1'
    ws['A1'].data_type = 'f'
    ws['B1'].value = '=C1+1'
    ws['B1'].data_type = 'f'
    ws['C1'].value = '=A1+1'
    ws['C1'].data_type = 'f'
    
    wb.save('indirect_circular.xlsx')
    print("Created indirect_circular.xlsx")

def create_valid_cross_reference():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Create valid cross-reference
    ws['A1'].value = 1  # Numeric value
    ws['A1'].data_type = 'n'
    ws['B1'].value = '=A1*2'  # Formula referencing A1
    ws['B1'].data_type = 'f'
    
    wb.save('valid_cross_reference.xlsx')
    print("Created valid_cross_reference.xlsx")

def verify_formulas():
    """Verify the formulas were saved correctly"""
    for filename in ['direct_circular.xlsx', 'indirect_circular.xlsx', 'valid_cross_reference.xlsx']:
        wb = openpyxl.load_workbook(f'{filename}', data_only=False)
        ws = wb.active
        print(f"\nVerifying {filename}:")
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith('='):
                    print(f"Cell {cell.coordinate}: Formula = {cell.value}")
                else:
                    print(f"Cell {cell.coordinate}: Value = {cell.value}")

def main():
    print("Generating Excel test files...")
    
    try:
        create_direct_circular_reference()
        create_indirect_circular_reference()
        create_valid_cross_reference()
        print("\nAll files generated successfully!")
        
        print("\nVerifying formula storage:")
        verify_formulas()
        
    except Exception as e:
        print(f"Error generating files: {str(e)}")

if __name__ == "__main__":
    main()