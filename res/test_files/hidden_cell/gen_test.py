import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension, RowDimension

def create_hidden_cells_test():
    wb = Workbook()
    ws = wb.active
    
    ws['A1'] = 'Visible'
    ws['B1'] = 'Hidden'
    
    ws.protection.sheet = True
    ws['B1'].protection = Protection(hidden=True)
    
    return wb

def create_hidden_rows_columns_test():
    wb = Workbook()
    ws = wb.active
    
    ws['A1'] = 'Content'
    ws['B2'] = 'More'
    ws['C3'] = 'Data'
    
    # Handle column dimensions
    col_dims = DimensionHolder(worksheet=ws)
    col_dims['A'] = ColumnDimension(ws, index='A', hidden=True)
    col_dims['C'] = ColumnDimension(ws, index='C', hidden=True)
    ws.column_dimensions = col_dims
    
    # Handle row dimensions separately
    for row in [1, 3]:
        ws.row_dimensions[row] = RowDimension(ws, index=row, hidden=True)
    
    return wb

def create_consecutive_hidden_rows_test():
    wb = Workbook()
    ws = wb.active
    
    for i in range(1, 6):
        ws[f'A{i}'] = f'Row {i}'
    
    for row in [1, 2, 4, 5]:
        ws.row_dimensions[row] = RowDimension(ws, index=row, hidden=True)
    
    return wb

def main():
    test_files = {
        'hidden_cells.xlsx': create_hidden_cells_test(),
        'hidden_rows_columns.xlsx': create_hidden_rows_columns_test(),
        'consecutive_hidden_rows.xlsx': create_consecutive_hidden_rows_test()
    }
    
    for filename, wb in test_files.items():
        wb.save(filename)
        print(f"Created {filename}")

if __name__ == "__main__":
    main()